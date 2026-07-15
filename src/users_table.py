from __future__ import annotations

import asyncio
import datetime
import json
import logging
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import List, Dict, Any, Optional, Set

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from rich.console import Console
from rich.table import Table
from aiohttp import web
import aiohttp
from classes import UserProfile
from client import Tuiclient


logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)-8s | %(message)s",
    datefmt="%H:%M:%S",
)
logger = logging.getLogger(__name__)



@dataclass
class ScrapingState:
    current_id: int = 0
    id_min: int = 9_900_000
    id_max: int = 15_000_000
    id_step: int = 1000
    total_ids: int = field(init=False)

    cooldown_seconds: float = 10.0
    start_time: float = field(default_factory=time.time)
    last_batch_time: float = field(default_factory=time.time)
    avg_batch_duration: float = 0.0
    batch_count: int = 0

    current_action: str = "idle"
    action_started_at: float = field(default_factory=time.time)
    action_duration_estimate: float = 0.0
    action_progress: float = 0.0
    action_detail: str = "" 

    is_running: bool = False
    is_paused: bool = False
    is_saving: bool = False
    last_error: Optional[str] = None

    reinit_logs: List[Dict[str, Any]] = field(default_factory=list)
    max_log_entries: int = 100

    users_scraped: int = 0
    excel_path: str = "users_table.xlsx"

    ws_clients: Set[web.WebSocketResponse] = field(default_factory=set)

    def __post_init__(self):
        self.total_ids = self.id_max - self.id_min

    def set_action(self, action: str, duration_estimate: float = 0.0, detail: str = "") -> None:
        """Set current action with optional duration estimate for progress bar."""
        self.current_action = action
        self.action_started_at = time.time()
        self.action_duration_estimate = max(duration_estimate, 0.001)
        self.action_progress = 0.0
        self.action_detail = detail

    def update_action_progress(self) -> None:
        """Update action_progress based on elapsed time vs estimate."""
        if self.action_duration_estimate <= 0:
            self.action_progress = 0.0
            return
        elapsed = time.time() - self.action_started_at
        self.action_progress = min(100.0, (elapsed / self.action_duration_estimate) * 100)

    @property
    def progress_percent(self) -> float:
        if self.total_ids <= 0:
            return 100.0
        done = self.current_id - self.id_min
        return min(100.0, (done / self.total_ids) * 100)

    @property
    def estimated_remaining_seconds(self) -> float:
        """ETA based on average batch duration."""
        if self.batch_count == 0 or not self.is_running:
            return 0.0
        remaining_ids = self.id_max - self.current_id
        remaining_batches = max(0, remaining_ids / self.id_step)
        return remaining_batches * (self.avg_batch_duration + self.cooldown_seconds)

    def to_dict(self) -> Dict[str, Any]:
        self.update_action_progress()
        return {
            "current_id": self.current_id,
            "id_min": self.id_min,
            "id_max": self.id_max,
            "progress_percent": round(self.progress_percent, 2),
            "cooldown_seconds": self.cooldown_seconds,
            "is_running": self.is_running,
            "is_paused": self.is_paused,
            "is_saving": self.is_saving,
            "users_scraped": self.users_scraped,
            "estimated_remaining_seconds": round(self.estimated_remaining_seconds, 1),
            "elapsed_seconds": round(time.time() - self.start_time, 1),
            "last_error": self.last_error,
            "reinit_logs": self.reinit_logs[-20:],
            "current_action": self.current_action,
            "action_progress": round(self.action_progress, 1),
            "action_detail": self.action_detail,
            "action_label": self._action_label(),
            "action_color": self._action_color(),
        }

    def _action_label(self) -> str:
        labels = {
            "idle": "Ожидание",
            "cooldown": "Кулдаун",
            "request": "Запрос к API",
            "parsing": "Парсинг данных",
            "saving": "Сохранение",
            "reinit": "Реинициализация",
        }
        return labels.get(self.current_action, self.current_action)

    def _action_color(self) -> str:
        colors = {
            "idle": "#8b949e",
            "cooldown": "#d29922",
            "request": "#58a6ff",
            "parsing": "#3fb950",
            "saving": "#a371f7",
            "reinit": "#f85149",
        }
        return colors.get(self.current_action, "#8b949e")

    def add_reinit_log(self, message: str) -> None:
        entry = {
            "timestamp": datetime.datetime.now().strftime("%H:%M:%S"),
            "message": message,
        }
        self.reinit_logs.append(entry)
        if len(self.reinit_logs) > self.max_log_entries:
            self.reinit_logs.pop(0)


def _init_worksheet(ws: Worksheet) -> None:
    ws.title = "Users"
    ws.views.sheetView[0].showGridLines = True

    title_font = Font(name="Segoe UI", size=16, bold=True)
    ws.merge_cells("A1:H1")
    ws["A1"] = "Список пользователей"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 40

    headers = ["ID", "Имена", "Телефон", "Регистрация", "URL", "Страна", "Статус", "Bio"]
    header_fill = PatternFill(start_color="00838F", end_color="00838F", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        if col_idx == 8:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        else:
            cell.alignment = header_alignment
    ws.row_dimensions[3].height = 28


def _append_user_row(ws: Worksheet, row_idx: int, user: UserProfile) -> None:
    thin_border = Border(
        left=Side(style="thin", color="E0E0E0"),
        right=Side(style="thin", color="E0E0E0"),
        top=Side(style="thin", color="E0E0E0"),
        bottom=Side(style="thin", color="E0E0E0"),
    )

    names_list = [f"{n.firstName} {n.lastName}".strip() for n in user.names]
    names_str = ", ".join(names_list) if names_list else "—"
    phone_str = f"+{user.phone}" if user.phone else "—"
    country_str = user.country if user.country else "—"
    reg_time_str = user.registrationTime.strftime("%Y-%m-%d %H:%M") if user.registrationTime else "—"
    bio_str = user.description if user.description else "—"
    url_str = user.baseUrl if user.baseUrl else "—"
    status_str = str(user.accountStatus) if user.accountStatus else "—"

    row_data = [
        user.id,
        names_str,
        phone_str,
        reg_time_str,
        url_str,
        country_str,
        status_str,
        bio_str,
    ]

    for col_idx, val in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        cell.border = thin_border
        cell.font = Font(name="Segoe UI", size=10, color="333333")
        if col_idx == 1:
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = "0"
        elif col_idx in (4, 6, 7):
            cell.alignment = Alignment(horizontal="center", vertical="center")
        elif col_idx == 3:
            cell.alignment = Alignment(horizontal="right", vertical="center")
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.row_dimensions[row_idx].height = 22


def _auto_size_columns(ws: Worksheet) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row < 3:
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

async def _broadcast_state(state: ScrapingState) -> None:
    if not state.ws_clients:
        return
    payload = json.dumps({"type": "state", "data": state.to_dict()})
    dead_clients: Set[web.WebSocketResponse] = set()

    for ws in state.ws_clients:
        try:
            await ws.send_str(payload)
        except Exception:
            dead_clients.add(ws)

    for ws in dead_clients:
        state.ws_clients.discard(ws)

async def _run_scraper(state: ScrapingState) -> None:
    state.is_running = True
    state.start_time = time.time()
    state.current_id = state.id_min

    cl = Tuiclient()
    await cl._init_log()
    await cl.connect()

    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("Failed to create worksheet")

    _init_worksheet(ws)
    current_row = 4

    console = Console()
    rich_table = Table(
        title="Users",
        title_style="bold magenta",
        show_header=True,
        header_style="bold cyan",
    )
    _init_rich_table(rich_table)

    try:
        while state.current_id < state.id_max and state.is_running:
            if state.is_paused:
                state.set_action("idle")
                await asyncio.sleep(0.5)
                await _broadcast_state(state)
                continue

            c_id_max = min(state.current_id + state.id_step, state.id_max)
            id_range = list(range(state.current_id, c_id_max))
            detail = f"IDs {state.current_id:,} — {c_id_max - 1:,}"

            logger.info("Fetching %s", detail)

            state.set_action("cooldown", duration_estimate=state.cooldown_seconds, detail=detail)
            cooldown_start = time.time()
            while time.time() - cooldown_start < state.cooldown_seconds:
                if not state.is_running or state.is_paused:
                    break
                await asyncio.sleep(0.1)
                await _broadcast_state(state)

            if not state.is_running or state.is_paused:
                continue

            state.set_action("request", duration_estimate=2.0, detail=detail)
            await _broadcast_state(state)

            batch_start = time.time()
            new_infos: List[UserProfile] = []
            try:
                new_infos = await cl.get_infos(id_range)
            except RuntimeError as exc:
                logger.error("RuntimeError during fetch: %s", exc)
                state.last_error = str(exc)
                state.add_reinit_log(f"RuntimeError: {exc}")

                await cl.disconnect()
                del cl
                logger.info("Reinitializing client...")
                state.add_reinit_log("Client reinitialized")

                reinit_duration = state.cooldown_seconds * 2
                state.set_action("reinit", duration_estimate=reinit_duration, detail="Подключение к API")
                reinit_start = time.time()
                while time.time() - reinit_start < reinit_duration:
                    if not state.is_running:
                        break
                    await asyncio.sleep(0.1)
                    await _broadcast_state(state)

                cl = Tuiclient()
                await cl._init_log()
                await cl.connect()
                continue

            batch_duration = time.time() - batch_start
            state.batch_count += 1
            state.avg_batch_duration = (
                state.avg_batch_duration * (state.batch_count - 1) + batch_duration
            ) / state.batch_count
            state.last_batch_time = time.time()

            users_by_id = {user.id: user for user in new_infos}

            parse_count = len([uid for uid in id_range if uid in users_by_id])
            state.set_action("parsing", duration_estimate=0.5 + parse_count * 0.01, detail=f"{parse_count} пользователей")
            await _broadcast_state(state)

            for uid in id_range:
                if uid not in users_by_id:
                    continue

                user = users_by_id[uid]
                state.users_scraped += 1

                if uid < state.id_min + 10:
                    try:
                        _add_rich_row(rich_table, user)
                    except Exception:
                        pass

                _append_user_row(ws, current_row, user)
                current_row += 1

                state.update_action_progress()
                if state.users_scraped % 10 == 0:
                    await _broadcast_state(state)

            state.current_id = c_id_max
            state.last_error = None

            if state.batch_count % 10 == 0:
                state.set_action("saving", duration_estimate=1.0, detail="Автосохранение")
                await _broadcast_state(state)
                _auto_size_columns(ws)
                wb.save(state.excel_path)
                logger.info("Auto-saved checkpoint: %s", state.excel_path)
                await _broadcast_state(state)

            await _broadcast_state(state)

        state.set_action("saving", duration_estimate=1.0, detail="Финальное сохранение")
        await _broadcast_state(state)
        _auto_size_columns(ws)
        wb.save(state.excel_path)
        logger.info("Final save: %s", state.excel_path)
        console.print(rich_table)

    finally:
        state.is_running = False
        state.set_action("idle")
        await cl.disconnect()
        await _broadcast_state(state)


def _init_rich_table(table: Table) -> None:
    table.add_column("ID", justify="right", style="dim")
    table.add_column("Имена", style="white")
    table.add_column("Телефон", justify="right")
    table.add_column("Регистрация", justify="center", style="blue")
    table.add_column("url", justify="center", style="blue")
    table.add_column("Страна", justify="center", style="yellow")
    table.add_column("Статус", justify="center")
    table.add_column("Bio", style="green")


def _add_rich_row(table: Table, user: UserProfile) -> None:
    names_list = [f"{n.firstName} {n.lastName}".strip() for n in user.names]
    names_str = ", ".join(names_list) if names_list else "—"
    phone_str = f"+{user.phone}" if user.phone else "—"
    country_str = user.country if user.country else "—"
    reg_time_str = user.registrationTime.strftime("%Y-%m-%d %H:%M") if user.registrationTime else "—"
    bio_str = user.description if user.description else "—"
    url_str = user.baseUrl if user.baseUrl else "—"
    status_str = str(user.accountStatus) if user.accountStatus else "—"

    table.add_row(
        str(user.id),
        names_str,
        phone_str,
        reg_time_str,
        url_str,
        country_str,
        status_str,
        bio_str,
    )


_DASHBOARD_HTML = """
<!DOCTYPE html>
<html lang="ru">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>User Scraper Dashboard</title>
    <script src="https://cdn.tailwindcss.com"></script>
    <script src="https://unpkg.com/lucide@latest"></script>
    <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js"></script>
    <script>
        tailwind.config = {
            theme: {
                extend: {
                    colors: {
                        bg: '#0f1117',
                        card: '#161b22',
                        border: '#30363d',
                        text: '#c9d1d9',
                        muted: '#8b949e',
                        accent: '#58a6ff',
                        success: '#3fb950',
                        warn: '#d29922',
                        danger: '#f85149',
                        purple: '#a371f7',
                    }
                }
            }
        }
    </script>
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
        body { font-family: 'Inter', system-ui, -apple-system, sans-serif; }

        ::-webkit-scrollbar { width: 6px; }
        ::-webkit-scrollbar-track { background: #161b22; }
        ::-webkit-scrollbar-thumb { background: #30363d; border-radius: 3px; }
        ::-webkit-scrollbar-thumb:hover { background: #484f58; }

        .progress-glow {
            box-shadow: 0 0 12px rgba(88, 166, 255, 0.4), 0 0 24px rgba(63, 185, 80, 0.2);
        }

        @keyframes pulse-dot {
            0%, 100% { opacity: 1; transform: scale(1); box-shadow: 0 0 0 0 currentColor; }
            50% { opacity: 0.6; transform: scale(0.85); box-shadow: 0 0 0 6px transparent; }
        }
        .animate-pulse-dot {
            animation: pulse-dot 1.5s ease-in-out infinite;
        }

        .card-hover {
            transition: transform 0.2s ease, box-shadow 0.2s ease;
        }
        .card-hover:hover {
            transform: translateY(-2px);
            box-shadow: 0 8px 32px rgba(0,0,0,0.3);
        }

        .btn-press:active {
            transform: scale(0.96);
        }

        .font-mono-nums {
            font-variant-numeric: tabular-nums;
            font-feature-settings: "tnum";
        }
    </style>
</head>
<body class="bg-bg text-text min-h-screen p-4 md:p-6 lg:p-8">
    <div class="max-w-7xl mx-auto">
        <div class="flex items-center gap-3 mb-6">
            <i data-lucide="search" class="w-7 h-7 text-accent"></i>
            <h1 class="text-2xl md:text-3xl font-semibold tracking-tight">User Scraper Dashboard</h1>
        </div>

        <div class="bg-card border border-border rounded-xl p-5 md:p-6 mb-5 card-hover">
            <div class="flex items-center justify-between mb-3">
                <div class="flex items-center gap-2">
                    <i data-lucide="bar-chart-3" class="w-5 h-5 text-accent"></i>
                    <span class="text-sm font-medium uppercase tracking-wider text-muted">Общий прогресс</span>
                </div>
                <span class="text-3xl md:text-4xl font-bold text-accent font-mono-nums" id="progress-text">0%</span>
            </div>
            <div class="w-full h-3 bg-border rounded-full overflow-hidden">
                <div id="progress-bar" class="h-full rounded-full transition-all duration-300 ease-out progress-glow"
                     style="width: 0%; background: linear-gradient(90deg, #58a6ff, #3fb950);"></div>
            </div>
            <div class="flex justify-between mt-2 text-sm text-muted">
                <span id="current-id">0</span>
                <span id="id-max">10,000,000</span>
            </div>
        </div>

        <div class="grid grid-cols-1 md:grid-cols-2 lg:grid-cols-4 gap-4 mb-5">
            <!-- Current Action Card -->
            <div id="action-card" class="bg-card border border-border rounded-xl p-5 card-hover hidden">
                <div class="flex items-center justify-between mb-3">
                    <div class="flex items-center gap-2">
                        <span id="action-dot" class="w-3 h-3 rounded-full animate-pulse-dot"></span>
                        <span id="action-label" class="text-sm font-medium uppercase tracking-wider text-muted">Ожидание</span>
                    </div>
                    <span id="action-percent" class="text-2xl font-bold font-mono-nums">0%</span>
                </div>
                <div class="w-full h-2 bg-border rounded-full overflow-hidden mb-2">
                    <div id="action-progress-fill" class="h-full rounded-full transition-all duration-300 ease-out"
                         style="width: 0%;"></div>
                </div>
                <p id="action-detail" class="text-sm text-muted truncate"></p>
            </div>

            <div class="bg-card border border-border rounded-xl p-5 card-hover">
                <div class="flex items-center gap-2 mb-3">
                    <i data-lucide="clock" class="w-5 h-5 text-warn"></i>
                    <span class="text-sm font-medium uppercase tracking-wider text-muted">Оставшееся время</span>
                </div>
                <div class="text-3xl font-bold text-warn font-mono-nums" id="eta">—</div>
                <p class="text-sm text-muted mt-2">Прошло: <span id="elapsed" class="text-text">0с</span></p>
            </div>

            <div class="bg-card border border-border rounded-xl p-5 card-hover">
                <div class="flex items-center gap-2 mb-3">
                    <i data-lucide="users" class="w-5 h-5 text-success"></i>
                    <span class="text-sm font-medium uppercase tracking-wider text-muted">Собрано пользователей</span>
                </div>
                <div class="text-3xl font-bold text-success font-mono-nums" id="users-scraped">0</div>
            </div>

            <div class="bg-card border border-border rounded-xl p-5 card-hover">
                <div class="flex items-center gap-2 mb-3">
                    <i data-lucide="activity" class="w-5 h-5 text-purple"></i>
                    <span class="text-sm font-medium uppercase tracking-wider text-muted">Статус</span>
                </div>
                <span id="status-badge" class="inline-flex items-center gap-1.5 px-3 py-1.5 rounded-full text-sm font-semibold bg-danger/15 text-danger">
                    <span class="w-2 h-2 rounded-full bg-danger"></span>
                    Остановлен
                </span>
                <p id="last-error" class="text-sm text-muted mt-2 truncate"></p>
            </div>
        </div>

        <div class="bg-card border border-border rounded-xl p-5 md:p-6 mb-5 card-hover">
            <div class="flex items-center gap-2 mb-4">
                <i data-lucide="trending-up" class="w-5 h-5 text-accent"></i>
                <span class="text-sm font-medium uppercase tracking-wider text-muted">Длительность итераций (сек)</span>
            </div>
            <div class="relative h-64 w-full">
                <canvas id="iteration-chart"></canvas>
            </div>
        </div>

        <div class="bg-card border border-border rounded-xl p-5 md:p-6 mb-5 card-hover">
            <div class="flex items-center gap-2 mb-4">
                <i data-lucide="sliders-horizontal" class="w-5 h-5 text-muted"></i>
                <span class="text-sm font-medium uppercase tracking-wider text-muted">Управление</span>
            </div>
            <div class="flex flex-wrap items-center gap-3">
                <label class="text-sm text-muted">Кулдаун (сек):</label>
                <input type="number" id="cooldown-input" value="10" min="0" step="0.5"
                       class="w-24 bg-bg border border-border rounded-lg px-3 py-2 text-sm text-text focus:outline-none focus:border-accent focus:ring-1 focus:ring-accent transition-colors">
                <button onclick="updateCooldown()" class="btn-press flex items-center gap-2 bg-accent hover:bg-accent/80 text-white px-4 py-2 rounded-lg text-sm font-medium transition-colors">
                    <i data-lucide="check" class="w-4 h-4"></i>
                    Применить
                </button>
                <button onclick="togglePause()" id="pause-btn" class="btn-press flex items-center gap-2 bg-border hover:bg-border/80 text-text px-4 py-2 rounded-lg text-sm font-medium transition-colors">
                    <i data-lucide="pause" class="w-4 h-4"></i>
                    <span>Пауза</span>
                </button>
                <button onclick="saveNow()" class="btn-press flex items-center gap-2 bg-success hover:bg-success/80 text-white px-4 py-2 rounded-lg text-sm font-medium transition-colors">
                    <i data-lucide="save" class="w-4 h-4"></i>
                    <span id="save-btn-text">Сохранить сейчас</span>
                </button>
                <button onclick="stopScraper()" class="btn-press flex items-center gap-2 bg-danger hover:bg-danger/80 text-white px-4 py-2 rounded-lg text-sm font-medium transition-colors">
                    <i data-lucide="square" class="w-4 h-4"></i>
                    Стоп
                </button>
            </div>
        </div>

        <div class="bg-card border border-border rounded-xl p-5 md:p-6 card-hover">
            <div class="flex items-center gap-2 mb-4">
                <i data-lucide="scroll-text" class="w-5 h-5 text-muted"></i>
                <span class="text-sm font-medium uppercase tracking-wider text-muted">Лог реинитов</span>
            </div>
            <div id="log-container" class="max-h-72 overflow-y-auto font-mono text-xs space-y-1">
                <div class="text-muted italic py-2">Лог пуст — пока нет реинитов</div>
            </div>
        </div>
    </div>

    <script>
        lucide.createIcons();

        const ctx = document.getElementById('iteration-chart').getContext('2d');
        const iterationChart = new Chart(ctx, {
            type: 'line',
            data: {
                labels: [],
                datasets: [{
                    label: 'Длительность (сек)',
                    data: [],
                    borderColor: '#58a6ff',
                    backgroundColor: 'rgba(88, 166, 255, 0.1)',
                    borderWidth: 2,
                    pointRadius: 3,
                    pointBackgroundColor: '#58a6ff',
                    pointBorderColor: '#0f1117',
                    pointBorderWidth: 2,
                    fill: true,
                    tension: 0.3
                }]
            },
            options: {
                responsive: true,
                maintainAspectRatio: false,
                interaction: {
                    mode: 'index',
                    intersect: false,
                },
                plugins: {
                    legend: { display: false },
                    tooltip: {
                        backgroundColor: '#161b22',
                        borderColor: '#30363d',
                        borderWidth: 1,
                        titleColor: '#8b949e',
                        bodyColor: '#c9d1d9',
                        padding: 10,
                        cornerRadius: 8,
                        displayColors: false,
                        callbacks: {
                            title: (items) => 'Итерация ' + items[0].label,
                            label: (item) => item.raw.toFixed(2) + ' сек'
                        }
                    }
                },
                scales: {
                    x: {
                        grid: { color: '#30363d', drawBorder: false },
                        ticks: { color: '#8b949e', font: { size: 11 } }
                    },
                    y: {
                        grid: { color: '#30363d', drawBorder: false },
                        ticks: { color: '#8b949e', font: { size: 11 } },
                        beginAtZero: true
                    }
                }
            }
        });

        const ws = new WebSocket(`ws://${location.host}/ws`);
        let isPaused = false;
        let iterationDurations = [];
        const MAX_CHART_POINTS = 50;

        function fmtTime(sec) {
            if (sec < 60) return Math.round(sec) + 'с';
            const m = Math.floor(sec / 60);
            const s = Math.round(sec % 60);
            if (m < 60) return m + 'м ' + s + 'с';
            const h = Math.floor(m / 60);
            return h + 'ч ' + (m % 60) + 'м ' + s + 'с';
        }

        function fmtNum(n) {
            return n.toLocaleString('ru-RU');
        }

        function updateChart(duration) {
            iterationDurations.push(duration);
            if (iterationDurations.length > MAX_CHART_POINTS) {
                iterationDurations.shift();
            }
            iterationChart.data.labels = iterationDurations.map((_, i) => i + 1);
            iterationChart.data.datasets[0].data = iterationDurations;
            iterationChart.update('none');
        }

        const actionColors = {
            cooldown: { bg: '#d29922', gradient: 'linear-gradient(90deg, #d29922, #e3b341)' },
            parsing: { bg: '#3fb950', gradient: 'linear-gradient(90deg, #3fb950, #56d364)' },
            request: { bg: '#58a6ff', gradient: 'linear-gradient(90deg, #58a6ff, #79c0ff)' },
            saving: { bg: '#a371f7', gradient: 'linear-gradient(90deg, #a371f7, #bc8cff)' },
            reinit: { bg: '#f85149', gradient: 'linear-gradient(90deg, #f85149, #ff7b72)' },
            idle: { bg: '#30363d', gradient: '#30363d' }
        };

        ws.onmessage = (event) => {
            try {
                const msg = JSON.parse(event.data);
                if (msg.type !== 'state') return;
                const d = msg.data;

                const progressText = document.getElementById('progress-text');
                const progressBar = document.getElementById('progress-bar');
                if (progressText) progressText.textContent = (d.progress_percent || 0) + '%';
                if (progressBar) progressBar.style.width = (d.progress_percent || 0) + '%';

                const currentIdEl = document.getElementById('current-id');
                const idMaxEl = document.getElementById('id-max');
                if (currentIdEl) currentIdEl.textContent = fmtNum(d.current_id || 0);
                if (idMaxEl) idMaxEl.textContent = fmtNum(d.id_max || 0);

                const etaEl = document.getElementById('eta');
                const elapsedEl = document.getElementById('elapsed');
                if (etaEl) etaEl.textContent = d.is_running ? fmtTime(d.estimated_remaining_seconds || 0) : '—';
                if (elapsedEl) elapsedEl.textContent = fmtTime(d.elapsed_seconds || 0);

                const usersScrapedEl = document.getElementById('users-scraped');
                if (usersScrapedEl) usersScrapedEl.textContent = fmtNum(d.users_scraped || 0);

                const badge = document.getElementById('status-badge');
                if (badge) {
                    if (d.is_running && !d.is_paused) {
                        badge.className = 'inline-flex items-center gap-1.5 px-3 py-1.5 rounded-full text-sm font-semibold bg-success/15 text-success';
                        badge.innerHTML = '<span class="w-2 h-2 rounded-full bg-success animate-pulse"></span>Выполняется';
                    } else if (d.is_paused) {
                        badge.className = 'inline-flex items-center gap-1.5 px-3 py-1.5 rounded-full text-sm font-semibold bg-warn/15 text-warn';
                        badge.innerHTML = '<span class="w-2 h-2 rounded-full bg-warn"></span>Пауза';
                    } else {
                        badge.className = 'inline-flex items-center gap-1.5 px-3 py-1.5 rounded-full text-sm font-semibold bg-danger/15 text-danger';
                        badge.innerHTML = '<span class="w-2 h-2 rounded-full bg-danger"></span>Остановлен';
                    }
                }

                const lastErrorEl = document.getElementById('last-error');
                if (lastErrorEl) lastErrorEl.textContent = d.last_error || '';

                const pauseBtn = document.getElementById('pause-btn');
                const pauseSpan = pauseBtn?.querySelector('span');
                const pauseIcon = pauseBtn?.querySelector('i');
                if (pauseBtn && pauseSpan) {
                    if (d.is_paused) {
                        pauseSpan.textContent = 'Продолжить';
                        if (pauseIcon) {
                            pauseIcon.setAttribute('data-lucide', 'play');
                            pauseIcon.classList.remove('lucide-pause');
                            pauseIcon.classList.add('lucide-play');
                        }
                    } else {
                        pauseSpan.textContent = 'Пауза';
                        if (pauseIcon) {
                            pauseIcon.setAttribute('data-lucide', 'pause');
                            pauseIcon.classList.remove('lucide-play');
                            pauseIcon.classList.add('lucide-pause');
                        }
                    }
                    lucide.createIcons();
                }
                isPaused = d.is_paused;

                const actionCard = document.getElementById('action-card');
                const actionLabel = document.getElementById('action-label');
                const actionDot = document.getElementById('action-dot');
                const actionDetail = document.getElementById('action-detail');
                const actionPercent = document.getElementById('action-percent');
                const actionFill = document.getElementById('action-progress-fill');

                if (d.is_running && d.current_action && d.current_action !== 'idle') {
                    if (actionCard) actionCard.classList.remove('hidden');
                    const color = d.action_color || '#8b949e';
                    const colors = actionColors[d.current_action] || actionColors.idle;

                    if (actionLabel) {
                        actionLabel.textContent = d.action_label || d.current_action;
                        actionLabel.style.color = color;
                    }
                    if (actionDot) {
                        actionDot.style.background = color;
                        actionDot.style.color = color;
                    }
                    if (actionDetail) actionDetail.textContent = d.action_detail || '';
                    if (actionPercent) {
                        actionPercent.textContent = Math.round(d.action_progress || 0) + '%';
                        actionPercent.style.color = color;
                    }
                    if (actionFill) {
                        actionFill.style.width = (d.action_progress || 0) + '%';
                        actionFill.style.background = colors.gradient;
                    }
                } else {
                    if (actionCard) actionCard.classList.add('hidden');
                }

                if (d.iteration_duration !== undefined && d.iteration_duration !== null) {
                    updateChart(d.iteration_duration);
                }

                const logContainer = document.getElementById('log-container');
                if (logContainer && d.reinit_logs && d.reinit_logs.length > 0) {
                    logContainer.innerHTML = d.reinit_logs.map((l, i) => {
                        let cls = 'text-text';
                        const msg = (l.message || '').toString();
                        if (msg.includes('Error') || msg.includes('RuntimeError')) cls = 'text-danger';
                        else if (msg.includes('reinit')) cls = 'text-warn';
                        return `<div class="log-entry-anim flex gap-3 py-1.5 border-b border-border/50 ${i === d.reinit_logs.length - 1 ? '' : ''}">
                            <span class="text-muted whitespace-nowrap shrink-0">${l.timestamp || ''}</span>
                            <span class="${cls}">${msg}</span>
                        </div>`;
                    }).join('');
                    logContainer.scrollTop = logContainer.scrollHeight;
                }
            } catch (err) {
                console.error('Dashboard update error:', err);
            }
        };

        ws.onopen = () => {
            console.log('[Dashboard] WebSocket connected');
        };

        ws.onerror = (err) => {
            console.error('[Dashboard] WebSocket error:', err);
        };

        ws.onclose = () => {
            console.log('[Dashboard] WebSocket closed');
            const badge = document.getElementById('status-badge');
            if (badge) {
                badge.className = 'inline-flex items-center gap-1.5 px-3 py-1.5 rounded-full text-sm font-semibold bg-danger/15 text-danger';
                badge.innerHTML = '<span class="w-2 h-2 rounded-full bg-danger"></span>Disconnected';
            }
            const actionCard = document.getElementById('action-card');
            if (actionCard) actionCard.classList.add('hidden');
        };

        async function updateCooldown() {
            const val = parseFloat(document.getElementById('cooldown-input').value);
            if (isNaN(val) || val < 0) return;
            await fetch('/api/cooldown', {
                method: 'POST',
                headers: {'Content-Type': 'application/json'},
                body: JSON.stringify({seconds: val})
            });
        }

        async function togglePause() {
            await fetch('/api/pause', {method: 'POST'});
        }

        async function saveNow() {
            const btnText = document.getElementById('save-btn-text');
            const btn = btnText?.parentElement;
            if (btn) btn.disabled = true;
            if (btnText) btnText.textContent = 'Сохраняю...';
            try {
                const res = await fetch('/api/save', {method: 'POST'});
                const data = await res.json();
                if (btnText) btnText.textContent = data.ok ? 'Сохранено!' : 'Ошибка';
            } catch {
                if (btnText) btnText.textContent = 'Ошибка';
            }
            setTimeout(() => {
                if (btnText) btnText.textContent = 'Сохранить сейчас';
                if (btn) btn.disabled = false;
            }, 2000);
        }

        async function stopScraper() {
            if (!confirm('Остановить скрапер? Прогресс сохранится.')) return;
            await fetch('/api/stop', {method: 'POST'});
        }
    </script>
</body>
</html>
"""


async def _handle_index(request: web.Request) -> web.Response:
    return web.Response(text=_DASHBOARD_HTML, content_type="text/html")


async def _handle_ws(request: web.Request) -> web.WebSocketResponse:
    state: ScrapingState = request.app["state"]
    ws = web.WebSocketResponse()
    await ws.prepare(request)
    state.ws_clients.add(ws)
    logger.info("WebSocket client connected: %s", request.remote)

    await ws.send_json({"type": "state", "data": state.to_dict()})

    try:
        async for msg in ws:
            if msg.type == aiohttp.WSMsgType.TEXT:
                pass
            elif msg.type == aiohttp.WSMsgType.ERROR:
                logger.error("WS error: %s", ws.exception())
    finally:
        state.ws_clients.discard(ws)
        logger.info("WebSocket client disconnected: %s", request.remote)

    return ws


async def _handle_cooldown(request: web.Request) -> web.Response:
    state: ScrapingState = request.app["state"]
    try:
        data = await request.json()
        seconds = float(data.get("seconds", 10))
        state.cooldown_seconds = max(0, seconds)
        logger.info("Cooldown updated to %.1f seconds", state.cooldown_seconds)
        await _broadcast_state(state)
        return web.json_response({"ok": True, "cooldown": state.cooldown_seconds})
    except Exception as exc:
        return web.json_response({"ok": False, "error": str(exc)}, status=400)


async def _handle_pause(request: web.Request) -> web.Response:
    state: ScrapingState = request.app["state"]
    state.is_paused = not state.is_paused
    logger.info("Scraper %s", "paused" if state.is_paused else "resumed")
    await _broadcast_state(state)
    return web.json_response({"ok": True, "paused": state.is_paused})


async def _handle_save(request: web.Request) -> web.Response:
    state: ScrapingState = request.app["state"]
    if state.is_saving:
        return web.json_response({"ok": False, "error": "Save already in progress"}, status=409)

    state.is_saving = True
    await _broadcast_state(state)

    try:
        path = Path(state.excel_path)
        if path.exists():
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            if ws:
                _auto_size_columns(ws)
                wb.save(path)
                logger.info("Manual save triggered: %s", path)
                return web.json_response({"ok": True, "path": str(path)})
        return web.json_response({"ok": False, "error": "Workbook not found"}, status=404)
    except Exception as exc:
        logger.error("Save failed: %s", exc)
        return web.json_response({"ok": False, "error": str(exc)}, status=500)
    finally:
        state.is_saving = False
        await _broadcast_state(state)


async def _handle_stop(request: web.Request) -> web.Response:
    state: ScrapingState = request.app["state"]
    state.is_running = False
    logger.info("Stop signal received")
    await _broadcast_state(state)
    return web.json_response({"ok": True})


def create_app(state: ScrapingState) -> web.Application:
    app = web.Application()
    app["state"] = state

    app.router.add_get("/", _handle_index)
    app.router.add_get("/ws", _handle_ws)
    app.router.add_post("/api/cooldown", _handle_cooldown)
    app.router.add_post("/api/pause", _handle_pause)
    app.router.add_post("/api/save", _handle_save)
    app.router.add_post("/api/stop", _handle_stop)

    return app


async def _periodic_broadcast(state: ScrapingState) -> None:
    while True:
        await asyncio.sleep(0.1)
        if state.is_running:
            await _broadcast_state(state)

async def main() -> None:
    state = ScrapingState(
        id_min=9_950_000,
        id_max=15_000_000,
        id_step=1000,
        cooldown_seconds=20.0,
    )

    app = create_app(state)

    scraper_task = asyncio.create_task(_run_scraper(state))
    broadcast_task = asyncio.create_task(_periodic_broadcast(state))

    runner = web.AppRunner(app)
    await runner.setup()
    site = web.TCPSite(runner, "127.0.0.1", 8081)
    await site.start()

    logger.info("🚀 Dashboard running at http://127.0.0.1:8081")

    try:
        await scraper_task
    except asyncio.CancelledError:
        pass
    finally:
        broadcast_task.cancel()
        await runner.cleanup()


if __name__ == "__main__":
    asyncio.run(main())
