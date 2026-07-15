import asyncio
import datetime
import sys
from typing import List, Optional
from pydantic import BaseModel

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from fastapi import FastAPI, BackgroundTasks
from fastapi.responses import HTMLResponse
import uvicorn

from rich.console import Console
from rich.table import Table

# Импорты ваших локальных модулей
from classes import UserProfile
from client import Tuiclient

# ==========================================
# ГЛОБАЛЬНОЕ СОСТОЯНИЕ ДЛЯ СИНХРОНИЗАЦИИ TUI/WEB
# ==========================================
class AppState:
    def __init__(self):
        # Настройки парсинга
        self.id_min: int = 9_990_000
        self.id_max: int = 10_000_000
        self.id_step: int = 1000
        self.current_id: int = self.id_min
        
        # Динамические настройки управления
        self.cooldown_seconds: float = 10.0  # Изменяется "на лету"
        
        # Статусы выполнения
        self.is_running: bool = False
        self.total_processed: int = 0
        self.logs: List[str] = []
        self.current_eta: str = "Расчет..."
        self.seconds_until_next_batch: float = 0.0
        
        # Флаги управления
        self.force_save_requested: bool = False
        self.stop_requested: bool = False
        
        # Данные для сохранения (кэш рабочей книги)
        self.wb: Optional[openpyxl.Workbook] = None
        self.ws: Optional[Worksheet] = None
        self.current_row: int = 4

    def log(self, message: str):
        timestamp = datetime.datetime.now().strftime("%H:%M:%S")
        self.logs.append(f"[{timestamp}] {message}")
        if len(self.logs) > 100:  # Ограничиваем размер лога в памяти
            self.logs.pop(0)

state = AppState()
app = FastAPI(title="Scraper Control Panel")
console = Console()

# ==========================================
# ШАБЛОНЫ EXCEL И RICH (БЕЗ ИЗМЕНЕНИЙ)
# ==========================================
def initrich(table: Table):
    table.add_column("ID", justify="right", style="dim")
    table.add_column("Имена", style="white")
    table.add_column("Телефон", justify="right")
    table.add_column("Регистрация", justify="center", style="blue")
    table.add_column("url", justify="center", style="blue")
    table.add_column("Страна", justify="center", style="yellow")
    table.add_column("Статус", justify="center")
    table.add_column("Bio", style="green")

def initws(ws: Worksheet):
    ws.title = "Users"
    ws.views.sheetView[0].showGridLines = True
    title_font = Font(name="Segoe UI", size=16, bold=True)
    ws.merge_cells("A1:H1")
    ws["A1"] = "Список пользователей"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 40

    headers = ["ID", "Имена", "Телефон", "Регистрация", "URL", "Страна", "Статус", "Bio"]
    header_fill = PatternFill(start_color="00838F", end_color="00838F")
    header_font = Font(name="Segoe UI", size=11, bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        if col_idx == 8:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        else:
            cell.alignment = header_alignment
    ws.row_dimensions[3].height = 28

def save_excel_file():
    """Безопасное сохранение текущего состояния Excel"""
    if not state.wb or not state.ws:
        return
    
    # Автоподгонка ширины колонок
    for col in state.ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.row < 3: 
                continue
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        state.ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
    
    filename = "users_table.xlsx"
    state.wb.save(filename)
    state.log(f"Таблица успешно сохранена в файл: {filename}")

# ==========================================
# ОСНОВНОЙ АСИНХРОННЫЙ ЦИКЛ ПАРСИНГА
# ==========================================
async def run_cooldown():
    """Динамический кулдаун с возможностью прерывания/изменения на лету"""
    start_time = asyncio.get_event_loop().time()
    while True:
        elapsed = asyncio.get_event_loop().time() - start_time
        remaining = max(0.0, state.cooldown_seconds - elapsed)
        state.seconds_until_next_batch = round(remaining, 1)
        
        if remaining <= 0:
            break
        await asyncio.sleep(0.1)

async def table_users_task():
    state.is_running = True
    cl = Tuiclient()
    await cl._init_log()
    await cl.connect()

    state.wb = openpyxl.Workbook()
    state.ws = state.wb.active
    if not state.ws:
        state.log("Ошибка инициализации Excel Workbook")
        state.is_running = False
        return
    
    table = Table(title="Users", title_style="bold magenta", show_header=True, header_style="bold cyan")
    initrich(table)
    initws(state.ws)

    thin_border = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )

    state.current_row = 4
    state.current_id = state.id_min
    start_time = datetime.datetime.now()

    while state.current_id < state.id_max and not state.stop_requested:
        c_id_max = min(state.current_id + state.id_step, state.id_max)

        # Динамический кулдаун перед пачкой данных
        await run_cooldown()

        # Проверка флага принудительного сохранения перед сетевым запросом
        if state.force_save_requested:
            save_excel_file()
            state.force_save_requested = False

        new_infos: List[UserProfile] = []
        try:
            target_ids = list(range(state.current_id, c_id_max))
            new_infos = await cl.get_infos(target_ids)
        except RuntimeError:
            state.log("RuntimeError во время запроса. Переинициализация клиента...")
            await cl.disconnect()
            state.log("Ожидание 100 секунд перед реинитом...")
            
            # Умный кулдаун реинита (можно прервать или отслеживать)
            reinit_seconds = 100
            for r_sec in range(reinit_seconds, 0, -1):
                state.seconds_until_next_batch = r_sec
                await asyncio.sleep(1)
                if state.force_save_requested:
                    save_excel_file()
                    state.force_save_requested = False

            cl = Tuiclient()
            await cl._init_log()
            await cl.connect()
            state.log("Реинициализация клиента завершена успешно.")
            continue  # Повторяем итерацию для текущего диапазона ID

        users_by_id = {user.id: user for user in new_infos}

        for i in range(state.current_id, c_id_max):
            if i not in users_by_id:
                continue
            
            user = users_by_id[i]
            names_list = [f"{n.firstName} {n.lastName}".strip() for n in user.names]
            names_str = ", ".join(names_list) if names_list else "—"
            phone_str = f"+{user.phone}" if user.phone else "—"
            country_str = user.country if user.country else "—"
            reg_time_str = user.registrationTime.strftime("%Y-%m-%d %H:%M") if user.registrationTime else "—"
            bio_str = user.description if user.description else "—"
            url_str = user.baseUrl if user.baseUrl else "—"
            status_str = str(user.accountStatus) if user.accountStatus else "—"

            # Вывод в консоль первых 10 записей сессии
            if i < state.id_min + 10:
                try:
                    table.add_row(
                        str(user.id), names_str, phone_str, reg_time_str,
                        url_str, country_str, status_str, bio_str
                    )
                except Exception:
                    pass

            # Запись строки в Excel
            row_data = [user.id, names_str, phone_str, reg_time_str, url_str, country_str, status_str, bio_str]
            for col_idx, val in enumerate(row_data, 1):
                cell = state.ws.cell(row=state.current_row, column=col_idx, value=val)
                cell.border = thin_border
                cell.font = Font(name="Segoe UI", size=10, color="333333")
                if col_idx == 1:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.number_format = "0"
                elif col_idx in (4, 6, 7):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                elif col_idx == 3:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

            state.ws.row_dimensions[state.current_row].height = 22
            state.current_row += 1
            state.total_processed += 1

        # Расчет ETA (Оставшееся время)
        total_to_process = state.id_max - state.id_min
        progress_pct = (state.total_processed / total_to_process) if total_to_process > 0 else 0
        
        elapsed_now = (datetime.datetime.now() - start_time).total_seconds()
        if progress_pct > 0:
            total_estimated_time = elapsed_now / progress_pct
            remaining_time = max(0.0, total_estimated_time - elapsed_now)
            
            hours, remainder = divmod(int(remaining_time), 3600)
            minutes, seconds = divmod(remainder, 60)
            state.current_eta = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
        else:
            state.current_eta = "Расчет..."

        state.current_id += state.id_step

    # Финал сессии
    await cl.disconnect()
    save_excel_file()
    console.print(table)
    state.is_running = False
    state.log("Парсинг успешно завершен!")


# ==========================================
# WEB API & ENDPOINTS (FASTAPI)
# ==========================================
class CooldownUpdate(BaseModel):
    seconds: float

@app.on_event("startup")
async def startup_event():
    # Запускаем основной цикл сбора данных асинхронно в фоне при старте веб-сервера
    asyncio.create_task(table_users_task())

@app.get("/api/status")
async def get_status():
    total_range = state.id_max - state.id_min
    progress_percentage = round((state.total_processed / total_range) * 100, 2) if total_range > 0 else 0
    return {
        "current_id": state.current_id,
        "total_processed": state.total_processed,
        "progress_percentage": progress_percentage,
        "cooldown_seconds": state.cooldown_seconds,
        "seconds_until_next_batch": state.seconds_until_next_batch,
        "eta": state.current_eta,
        "is_running": state.is_running,
        "logs": state.logs[-20:]  # Отдаем только последние 20 логов на веб
    }

@app.post("/api/cooldown")
async def update_cooldown(data: CooldownUpdate):
    if data.seconds < 0.1:
        return {"status": "error", "message": "Интервал не может быть меньше 0.1 сек."}
    state.cooldown_seconds = data.seconds
    state.log(f"Длительность кулдауна изменена на {data.seconds} сек.")
    return {"status": "success", "cooldown_seconds": state.cooldown_seconds}

@app.post("/api/save")
async def trigger_save():
    if not state.is_running:
        return {"status": "error", "message": "Скрипт сейчас не работает."}
    state.force_save_requested = True
    state.log("Запрошено принудительное сохранение таблицы...")
    return {"status": "success", "message": "Файл будет сохранен в ближайшие секунды."}

@app.get("/", response_class=HTMLResponse)
async def index():
    html_content = """
    <!DOCTYPE html>
    <html lang="ru">
    <head>
        <meta charset="UTF-8">
        <title>Parser Dashboard</title>
        <script src="https://cdn.tailwindcss.com"></script>
    </head>
    <body class="bg-gray-900 text-gray-100 font-sans min-h-screen">
        <div class="max-w-4xl mx-auto py-8 px-4">
            <header class="flex justify-between items-center mb-8 border-b border-gray-800 pb-4">
                <div>
                    <h1 class="text-3xl font-bold text-cyan-400">Scraper Control Center</h1>
                    <p class="text-gray-400 text-sm mt-1">Управление фоновым парсингом в реальном времени</p>
                </div>
                <div class="flex items-center space-x-2">
                    <span class="relative flex h-3 w-3">
                        <span id="status-ring" class="animate-ping absolute inline-flex h-full w-full rounded-full bg-green-400 opacity-75"></span>
                        <span id="status-dot" class="relative inline-flex rounded-full h-3 w-3 bg-green-500"></span>
                    </span>
                    <span id="status-text" class="text-sm font-semibold">Активен</span>
                </div>
            </header>

            <div class="grid grid-cols-1 md:grid-cols-3 gap-6 mb-8">
                <div class="bg-gray-800 p-6 rounded-lg border border-gray-700">
                    <h3 class="text-gray-400 text-xs uppercase tracking-wider font-semibold">Общий Прогресс</h3>
                    <p class="text-3xl font-extrabold mt-2 text-white" id="progress-text">0%</p>
                    <div class="w-full bg-gray-700 h-2.5 rounded-full mt-4 overflow-hidden">
                        <div id="progress-bar" class="bg-cyan-500 h-2.5 rounded-full transition-all duration-500" style="width: 0%"></div>
                    </div>
                </div>

                <div class="bg-gray-800 p-6 rounded-lg border border-gray-700">
                    <h3 class="text-gray-400 text-xs uppercase tracking-wider font-semibold">Следующая итерация</h3>
                    <p class="text-3xl font-extrabold mt-2 text-yellow-400" id="timer-text">0.0s</p>
                    <p class="text-xs text-gray-400 mt-2">Базовый кулдаун: <span id="base-cooldown">0</span>с</p>
                </div>

                <div class="bg-gray-800 p-6 rounded-lg border border-gray-700">
                    <h3 class="text-gray-400 text-xs uppercase tracking-wider font-semibold">Оставшееся время (ETA)</h3>
                    <p class="text-3xl font-extrabold mt-2 text-green-400" id="eta-text">Расчет...</p>
                    <p class="text-xs text-gray-400 mt-2">Обработано: <span id="processed-count">0</span> лимитов</p>
                </div>
            </div>

            <div class="grid grid-cols-1 md:grid-cols-2 gap-6 mb-8">
                <div class="bg-gray-800 p-6 rounded-lg border border-gray-700">
                    <h2 class="text-lg font-semibold mb-4 text-white">Управление задержкой</h2>
                    <div class="flex items-center space-x-4">
                        <input type="number" id="cooldown-input" min="0.5" step="0.5" class="bg-gray-700 text-white rounded px-4 py-2 w-32 focus:outline-none focus:ring-2 focus:ring-cyan-500" />
                        <button onclick="updateCooldown()" class="bg-cyan-600 hover:bg-cyan-500 text-white font-bold py-2 px-6 rounded transition duration-200">
                            Применить
                        </button>
                    </div>
                    <p class="text-xs text-gray-400 mt-2">Изменяет интервал между пакетами ID прямо на лету.</p>
                </div>

                <div class="bg-gray-800 p-6 rounded-lg border border-gray-700 flex flex-col justify-between">
                    <div>
                        <h2 class="text-lg font-semibold mb-2 text-white">Сохранить состояние</h2>
                        <p class="text-xs text-gray-400">Нажмите кнопку, чтобы принудительно выгрузить текущие собранные данные в <code>users_table.xlsx</code> без прерывания скрипта.</p>
                    </div>
                    <button onclick="triggerSave()" class="bg-emerald-600 hover:bg-emerald-500 text-white font-bold py-2 px-6 rounded transition duration-200 mt-4">
                        Сохранить Excel сейчас
                    </button>
                </div>
            </div>

            <div class="bg-gray-800 p-6 rounded-lg border border-gray-700">
                <h2 class="text-lg font-semibold mb-4 text-white">События и Реинициализации</h2>
                <div id="log-container" class="bg-gray-950 p-4 rounded h-48 overflow-y-auto font-mono text-xs text-green-400 space-y-1">
                    </div>
            </div>
        </div>

        <script>
            async function updateStatus() {
                try {
                    const response = await fetch('/api/status');
                    const data = await response.json();

                    // Обновляем прогресс-бар
                    document.getElementById('progress-text').innerText = data.progress_percentage + '%';
                    document.getElementById('progress-bar').style.width = data.progress_percentage + '%';

                    // Таймер до следующего пакета
                    document.getElementById('timer-text').innerText = data.seconds_until_next_batch + 's';
                    document.getElementById('base-cooldown').innerText = data.cooldown_seconds;

                    // Общие счетчики и ETA
                    document.getElementById('eta-text').innerText = data.eta;
                    document.getElementById('processed-count').innerText = data.total_processed;

                    // Установка input поля, если фокус на нем отсутствует
                    const input = document.getElementById('cooldown-input');
                    if (document.activeElement !== input && input.value !== String(data.cooldown_seconds)) {
                        input.value = data.cooldown_seconds;
                    }

                    // Индикатор работы
                    const ring = document.getElementById('status-ring');
                    const dot = document.getElementById('status-dot');
                    const text = document.getElementById('status-text');
                    if (data.is_running) {
                        ring.className = "animate-ping absolute inline-flex h-full w-full rounded-full bg-green-400 opacity-75";
                        dot.className = "relative inline-flex rounded-full h-3 w-3 bg-green-500";
                        text.innerText = "Активен";
                    } else {
                        ring.className = "hidden";
                        dot.className = "relative inline-flex rounded-full h-3 w-3 bg-red-500";
                        text.innerText = "Остановлен";
                    }

                    // Логи
                    const logContainer = document.getElementById('log-container');
                    logContainer.innerHTML = data.logs.map(log => `<div>${log}</div>`).join('');
                } catch (err) {
                    console.error("Ошибка обновления статуса:", err);
                }
            }

            async function updateCooldown() {
                const seconds = parseFloat(document.getElementById('cooldown-input').value);
                await fetch('/api/cooldown', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify({ seconds: seconds })
                });
            }

            async function triggerSave() {
                const response = await fetch('/api/save', { method: 'POST' });
                const result = await response.json();
                alert(result.message);
            }

            // Опрашиваем бэкенд раз в секунду
            setInterval(updateStatus, 1000);
            updateStatus();
        </script>
    </body>
    </html>
    """
    return HTMLResponse(content=html_content)


def main():
    uvicorn.run(app, host="127.0.0.1", port=8000, log_level="warning")

if __name__ == "__main__":
    main()